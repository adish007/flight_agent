# main.py
import argparse
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta

import pandas as pd

from config import (
    ORIGIN,
    DESTINATIONS,
    NUM_ADULTS,
    MAX_DURATION_HOURS,
    MAX_RETRIES,
    MAX_WORKERS,
    TRIP_LENGTHS,
    generate_dates,
)
from scraper import search_flights, random_delay

RUNS_DIR = "runs"

# File names (will be prefixed with run directory in main())
LEGS_FILE = "legs.csv"
OUTPUT_FILE_FILTERED = "flights_filtered.csv"
ERRORS_LOG = "errors.log"
PROGRESS_FILE = "progress.json"
LEGS_COLUMNS = [
    "direction",  # "outbound" or "return"
    "destination",
    "city_name",
    "date",
    "price",
    "airline",
    "duration_hrs",
    "stops",
    "departure_time",
    "arrival_time",
]


def load_progress() -> set:
    """Load set of already-scraped keys."""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            return set(tuple(x) for x in json.load(f))
    return set()


def save_progress(completed: set) -> None:
    """Save progress to disk."""
    with open(PROGRESS_FILE, "w") as f:
        json.dump(list(completed), f)


def log_error(msg: str) -> None:
    """Append error to log file."""
    with open(ERRORS_LOG, "a") as f:
        f.write(f"{msg}\n")


def parse_duration_hrs(duration_str: str) -> float | None:
    """Parse duration string like '6 hr 25 min' to decimal hours."""
    if not duration_str:
        return None
    match_hr = re.search(r"(\d+)\s*hr", duration_str)
    hours = int(match_hr.group(1)) if match_hr else 0
    match_min = re.search(r"(\d+)\s*min", duration_str)
    minutes = int(match_min.group(1)) if match_min else 0
    return hours + minutes / 60


def parse_price(price_str: str) -> int | None:
    """Parse price string like '$507' to integer."""
    if not price_str:
        return None
    match = re.search(r"\$?([\d,]+)", price_str)
    if match:
        return int(match.group(1).replace(",", ""))
    return None


def parse_stops(stops_val) -> int:
    """Parse stops value (str or int) to integer."""
    if isinstance(stops_val, int):
        return stops_val
    if not stops_val or stops_val == "Unknown":
        return -1
    if "nonstop" in str(stops_val).lower():
        return 0
    match = re.search(r"(\d+)", str(stops_val))
    return int(match.group(1)) if match else -1


def search_and_save_leg(
    direction: str, origin: str, dest: str, dest_code: str, city_name: str, date_str: str
) -> list[dict]:
    """Search one leg and return parsed results ready for CSV."""
    for attempt in range(MAX_RETRIES):
        raw = search_flights(origin, dest, date_str, NUM_ADULTS, max_stops=1)
        if raw:
            break
        if attempt < MAX_RETRIES - 1:
            random_delay()
    else:
        return []

    results = []
    seen = set()
    for f in raw:
        dur = parse_duration_hrs(f.get("duration", ""))
        price = parse_price(f.get("price", ""))
        if dur is None or price is None:
            continue
        if dur >= MAX_DURATION_HOURS:
            continue

        key = (f.get("airline"), f.get("departure_time"), price)
        if key in seen:
            continue
        seen.add(key)

        results.append({
            "direction": direction,
            "destination": dest_code,
            "city_name": city_name,
            "date": date_str,
            "price": price,
            "airline": f.get("airline", ""),
            "duration_hrs": round(dur, 2),
            "stops": parse_stops(f.get("stops", "")),
            "departure_time": f.get("departure_time", ""),
            "arrival_time": f.get("arrival_time", ""),
        })

    return results


def append_legs_csv(legs: list[dict]) -> None:
    """Append leg records to the legs CSV file."""
    import csv

    file_exists = os.path.exists(LEGS_FILE)
    with open(LEGS_FILE, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=LEGS_COLUMNS, extrasaction="ignore")
        if not file_exists:
            writer.writeheader()
        writer.writerows(legs)


EXCEL_REPORT_FILE = "flights_report.xlsx"  # set to run dir in main()
BUDGET_AIRLINES = ["Frontier", "Spirit"]
TOP_N = 5

REPORT_COLUMNS = {
    "depart_date": "Depart",
    "return_date": "Return",
    "trip_days": "Days",
    "total_price": "Total Price",
    "outbound_airline": "Outbound Airline",
    "return_airline": "Return Airline",
    "outbound_duration_hrs": "Out Duration (hrs)",
    "return_duration_hrs": "Ret Duration (hrs)",
    "outbound_stops": "Out Stops",
    "return_stops": "Ret Stops",
}
# Google Flights link column is appended after REPORT_COLUMNS


def _build_google_flights_url(origin: str, dest: str, depart_date: str, return_date: str) -> str:
    """Build a Google Flights search URL using the same encoding as fast-flights."""
    from fast_flights import FlightData, Passengers
    from fast_flights.flights_impl import TFSData

    tfs = TFSData.from_interface(
        flight_data=[
            FlightData(date=depart_date, from_airport=origin, to_airport=dest),
            FlightData(date=return_date, from_airport=dest, to_airport=origin),
        ],
        trip="round-trip",
        passengers=Passengers(adults=NUM_ADULTS),
        seat="economy",
    )
    encoded = tfs.as_b64().decode("utf-8")
    return f"https://www.google.com/travel/flights?tfs={encoded}&hl=en&curr=USD"


def _write_excel_sheet(ws, trips_df: pd.DataFrame) -> None:
    """Write grouped top-5-per-destination data to an openpyxl worksheet."""
    from openpyxl.styles import Font

    bold = Font(bold=True)
    header_font = Font(bold=True, size=13)
    link_font = Font(color="0563C1", underline="single")
    col_headers = list(REPORT_COLUMNS.values()) + ["Google Flights"]
    link_col = len(REPORT_COLUMNS) + 1

    row = 1
    # Sort destinations alphabetically by city name
    for dest, group in trips_df.groupby("city_name", sort=True):
        dest_code = group.iloc[0]["destination"]
        top = group.nsmallest(TOP_N, "total_price")

        # Destination header
        ws.cell(row=row, column=1, value=f"{dest} ({dest_code})").font = header_font
        row += 1

        # Column headers
        for col_idx, header in enumerate(col_headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = bold
        row += 1

        # Data rows
        for _, trip in top.iterrows():
            for col_idx, src_col in enumerate(REPORT_COLUMNS.keys(), start=1):
                val = trip[src_col]
                cell = ws.cell(row=row, column=col_idx, value=val)
                if src_col == "total_price":
                    cell.number_format = "$#,##0"

            # Google Flights hyperlink
            url = _build_google_flights_url(
                ORIGIN, trip["destination"], trip["depart_date"], trip["return_date"]
            )
            cell = ws.cell(row=row, column=link_col, value="Search")
            cell.hyperlink = url
            cell.font = link_font

            row += 1

        # Blank separator row
        row += 1

    # Auto-fit column widths
    for col_idx in range(1, len(col_headers) + 1):
        max_len = 0
        for r in range(1, row):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3


def build_excel_report(trips_df: pd.DataFrame, legs_df: pd.DataFrame) -> None:
    """Build a formatted Excel report with two tabs from round-trip data."""
    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: All Airlines
    ws_all = wb.active
    ws_all.title = "All Airlines"
    _write_excel_sheet(ws_all, trips_df)

    # Sheet 2: No Budget Airlines — filter budget airlines from raw legs FIRST,
    # then rebuild round trips so non-budget options surface for every destination
    budget_pattern = "|".join(BUDGET_AIRLINES)
    non_budget_legs = legs_df[
        ~legs_df["airline"].str.contains(budget_pattern, case=False, na=False)
    ]
    no_budget_trips = _combine_legs_into_trips(non_budget_legs)
    ws_no_budget = wb.create_sheet("No Budget Airlines")
    _write_excel_sheet(ws_no_budget, no_budget_trips)

    wb.save(EXCEL_REPORT_FILE)
    print(f"Excel report saved to {EXCEL_REPORT_FILE}")


def _combine_legs_into_trips(legs_df: pd.DataFrame) -> pd.DataFrame:
    """Combine outbound + return legs into round trips for each trip length.

    Returns a DataFrame of round-trip combinations sorted by total_price.
    """
    outbound = legs_df[legs_df["direction"] == "outbound"].copy()
    returns = legs_df[legs_df["direction"] == "return"].copy()

    if outbound.empty or returns.empty:
        return pd.DataFrame()

    best_outbound = (
        outbound.sort_values("price")
        .groupby(["destination", "date"])
        .first()
        .reset_index()
    )
    best_return = (
        returns.sort_values("price")
        .groupby(["destination", "date"])
        .first()
        .reset_index()
    )

    trips = []
    for _, out_row in best_outbound.iterrows():
        depart_date = date.fromisoformat(out_row["date"])
        dest = out_row["destination"]

        for trip_days in TRIP_LENGTHS:
            return_date = depart_date + timedelta(days=trip_days)
            return_date_str = return_date.strftime("%Y-%m-%d")

            ret_match = best_return[
                (best_return["destination"] == dest)
                & (best_return["date"] == return_date_str)
            ]
            if ret_match.empty:
                continue

            ret_row = ret_match.iloc[0]
            total = int(out_row["price"]) + int(ret_row["price"])

            trips.append({
                "destination": dest,
                "city_name": out_row["city_name"],
                "depart_date": out_row["date"],
                "return_date": return_date_str,
                "trip_days": trip_days,
                "outbound_price": int(out_row["price"]),
                "return_price": int(ret_row["price"]),
                "total_price": total,
                "outbound_airline": out_row["airline"],
                "return_airline": ret_row["airline"],
                "outbound_duration_hrs": out_row["duration_hrs"],
                "return_duration_hrs": ret_row["duration_hrs"],
                "outbound_stops": int(out_row["stops"]),
                "return_stops": int(ret_row["stops"]),
            })

    if not trips:
        return pd.DataFrame()

    return pd.DataFrame(trips).sort_values("total_price")


def build_round_trips() -> None:
    """Post-process: combine outbound + return legs into round trips for each trip length."""
    if not os.path.exists(LEGS_FILE):
        print("No legs data found.")
        return

    df = pd.read_csv(LEGS_FILE)

    trips_df = _combine_legs_into_trips(df)
    if trips_df.empty:
        print("No valid round trips found.")
        return

    trips_df.to_csv(OUTPUT_FILE_FILTERED, index=False)

    print(f"\nGenerated {len(trips_df)} round-trip combinations.")
    print(f"Saved to {OUTPUT_FILE_FILTERED}")
    print(f"\nTop 20 cheapest round trips:")
    cols = [
        "destination", "city_name", "depart_date", "return_date",
        "trip_days", "total_price", "outbound_airline", "return_airline",
    ]
    print(trips_df[cols].head(20).to_string(index=False))

    # Build formatted Excel report — pass raw legs so non-budget tab
    # can rebuild trips from non-budget legs instead of just filtering trips
    build_excel_report(trips_df, df)


def parse_args():
    parser = argparse.ArgumentParser(description="Search Google Flights for cheap Caribbean round trips.")
    parser.add_argument(
        "--start-date", required=True,
        help="Start date for search range (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--end-date", required=True,
        help="End date for search range (YYYY-MM-DD)",
    )
    args = parser.parse_args()
    try:
        start = date.fromisoformat(args.start_date)
        end = date.fromisoformat(args.end_date)
    except ValueError:
        parser.error("Dates must be in YYYY-MM-DD format")
    if start > end:
        parser.error("--start-date must be before --end-date")
    return start, end


def _ensure_playwright():
    """Auto-install Chromium for Playwright if not already installed."""
    import subprocess
    import shutil
    # Check if chromium is already installed by looking for the browser
    try:
        from playwright._impl._driver import compute_driver_executable
        driver = compute_driver_executable()
        result = subprocess.run(
            [str(driver), "install", "--dry-run", "chromium"],
            capture_output=True, text=True
        )
        if "chromium" in result.stdout.lower():
            print("Installing Chromium browser (one-time setup)...")
            subprocess.run([str(driver), "install", "chromium"], check=True)
    except Exception:
        # Fallback: just run playwright install chromium
        playwright_cmd = shutil.which("playwright")
        if playwright_cmd:
            subprocess.run([playwright_cmd, "install", "chromium"], check=True)


def main():
    global LEGS_FILE, OUTPUT_FILE_FILTERED, ERRORS_LOG, PROGRESS_FILE, EXCEL_REPORT_FILE

    _ensure_playwright()
    start, end = parse_args()

    # Create run directory: runs/<today>/run_N/
    today = date.today().isoformat()
    day_dir = os.path.join(RUNS_DIR, today)
    os.makedirs(day_dir, exist_ok=True)

    # Find next run number
    existing = [d for d in os.listdir(day_dir) if d.startswith("run_") and os.path.isdir(os.path.join(day_dir, d))]
    next_num = max((int(d.split("_")[1]) for d in existing), default=0) + 1
    run_dir = os.path.join(day_dir, f"run_{next_num}")
    os.makedirs(run_dir)

    # Point all output files into the run directory
    LEGS_FILE = os.path.join(run_dir, "legs.csv")
    OUTPUT_FILE_FILTERED = os.path.join(run_dir, "flights_filtered.csv")
    ERRORS_LOG = os.path.join(run_dir, "errors.log")
    PROGRESS_FILE = os.path.join(run_dir, "progress.json")
    EXCEL_REPORT_FILE = os.path.join(run_dir, "flights_report.xlsx")

    dates = generate_dates(start, end)
    completed = load_progress()
    destinations = list(DESTINATIONS.items())

    # Build search tasks: one outbound + one return per (destination, date)
    tasks = []
    for dest_code, city_name in destinations:
        for search_date in dates:
            # Outbound: BOS -> dest
            key_out = ("outbound", dest_code, search_date)
            if key_out not in completed:
                tasks.append((key_out, ORIGIN, dest_code, dest_code, city_name, search_date))
            # Return: dest -> BOS
            key_ret = ("return", dest_code, search_date)
            if key_ret not in completed:
                tasks.append((key_ret, dest_code, ORIGIN, dest_code, city_name, search_date))

    # Also search return dates that extend into June (for trips departing late May)
    max_trip = max(TRIP_LENGTHS)
    extra_end = date.fromisoformat(dates[-1]) + timedelta(days=max_trip)
    extra_start = date.fromisoformat(dates[-1]) + timedelta(days=1)
    extra_date = extra_start
    while extra_date <= extra_end:
        date_str = extra_date.strftime("%Y-%m-%d")
        for dest_code, city_name in destinations:
            key_ret = ("return", dest_code, date_str)
            if key_ret not in completed:
                tasks.append((key_ret, dest_code, ORIGIN, dest_code, city_name, date_str))
        extra_date += timedelta(days=1)

    total_legs = len(destinations) * len(dates) * 2
    already_done = len(completed)

    print("Caribbean Flight Search Agent (via Google Flights)")
    print(f"Run output: {run_dir}")
    print(f"Origin: {ORIGIN}")
    print(f"Destinations: {len(destinations)}")
    print(f"Dates: {dates[0]} to {dates[-1]} ({len(dates)} days)")
    print(f"Trip lengths: {TRIP_LENGTHS} days (computed in post-processing)")
    print(f"Searches: {len(tasks)} remaining | {already_done} already done")
    print(f"Parallel workers: {MAX_WORKERS}")
    print("---")

    if not tasks:
        print("All searches already complete! Running post-processing...")
        build_round_trips()
        return

    count = 0

    try:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_task = {}
            for key, origin, dest, dest_code, city_name, search_date in tasks:
                direction = key[0]
                future = executor.submit(
                    search_and_save_leg, direction, origin, dest, dest_code, city_name, search_date
                )
                future_to_task[future] = key

            for future in as_completed(future_to_task):
                key = future_to_task[future]
                direction, dest_code, search_date = key
                count += 1

                try:
                    legs = future.result()
                except Exception as e:
                    log_error(f"{direction},{dest_code},{search_date},{e}")
                    print(f"  [{count}/{len(tasks)}] {direction} {dest_code} {search_date} ERROR: {e}")
                    completed.add(key)
                    if count % 20 == 0:
                        save_progress(completed)
                    continue

                if legs:
                    append_legs_csv(legs)
                    cheapest = min(legs, key=lambda x: x["price"])
                    print(
                        f"  [{count}/{len(tasks)}] {direction:8s} {dest_code} {search_date} "
                        f"— {len(legs)} flights, best ${cheapest['price']} ({cheapest['airline']})"
                    )
                else:
                    print(f"  [{count}/{len(tasks)}] {direction:8s} {dest_code} {search_date} — no flights under {MAX_DURATION_HOURS}h")

                completed.add(key)
                if count % 20 == 0:
                    save_progress(completed)

    except KeyboardInterrupt:
        print(f"\n\nInterrupted! Progress saved. {len(completed)} searches completed.")

    save_progress(completed)

    # Post-processing: build round trips from legs
    print(f"\n{'='*60}")
    print("Post-processing: building round-trip combinations...")
    build_round_trips()


if __name__ == "__main__":
    main()
